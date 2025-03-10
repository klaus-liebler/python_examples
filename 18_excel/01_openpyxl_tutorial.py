from openpyxl import Workbook
wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("AmEnde") # insert at the end (default)
ws2 = wb.create_sheet("AnErsterPosition", 0) # insert at first position
ws3 = wb.create_sheet("AnVorletzterPosition", -1) # insert at the penultimate position
ws.title = "NeuerName"
ws_alternativer_zugriff = wb["NeuerName"]
print(wb.sheetnames)
ws['A4'] = 42
c = ws['A4']
print(c.value)
cell_range = ws['A1':'C2']
colC = ws['C']
col_range = ws['C:D']
row10 = ws[10]
row_range = ws[5:10]
print("\n\n=====ws.iter_rows")
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2):
   for cell in row:
       print(cell)
print("\n\n=====ws.iter_cols")
for col in ws.iter_cols(min_row=1, max_col=3, max_row=2):
    for cell in col:
        print(cell)

wb.save('excel_file.xlsx')